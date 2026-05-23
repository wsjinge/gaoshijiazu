#!/usr/bin/env python3
import openpyxl
import json
import sys

def convert_xls_to_json(xls_file):
    try:
        # 打开 Excel 文件（openpyxl 主要支持 .xlsx，但也能处理某些 .xls）
        workbook = openpyxl.load_workbook(xls_file, data_only=True)
        print(f'工作簿包含 {len(workbook.sheetnames)} 个工作表')
        print(f'工作表名称: {workbook.sheetnames}')

        # 转换所有工作表
        all_data = {}
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            print(f'\n处理工作表 "{sheet_name}": {sheet.max_row} 行, {sheet.max_column} 列')

            # 获取标题行（假设第一行是标题）
            if sheet.max_row > 0:
                headers = []
                for j in range(1, sheet.max_column + 1):
                    cell_value = sheet.cell(1, j).value
                    headers.append(str(cell_value) if cell_value is not None else f'Column_{j}')
                print(f'标题: {headers}')

                # 读取数据行
                data_rows = []
                for i in range(2, sheet.max_row + 1):
                    row_data = {}
                    has_data = False
                    for j in range(1, sheet.max_column + 1):
                        cell = sheet.cell(i, j)
                        cell_value = cell.value

                        if cell_value is not None:
                            has_data = True
                            # 处理不同类型的单元格值
                            if isinstance(cell_value, str):
                                row_data[headers[j-1]] = cell_value
                            elif isinstance(cell_value, (int, float)):
                                row_data[headers[j-1]] = cell_value
                            elif isinstance(cell_value, bool):
                                row_data[headers[j-1]] = cell_value
                            else:
                                row_data[headers[j-1]] = str(cell_value)
                        else:
                            row_data[headers[j-1]] = None

                    if has_data:  # 只添加非空行
                        data_rows.append(row_data)

                all_data[sheet_name] = {
                    'headers': headers,
                    'rows': data_rows,
                    'total': len(data_rows)
                }

        return all_data

    except Exception as e:
        print(f'错误: {e}')
        import traceback
        traceback.print_exc()
        return None

if __name__ == '__main__':
    xls_file = '高氏家族.xls'
    data = convert_xls_to_json(xls_file)

    if data:
        # 输出 JSON
        output_file = '高氏家族.json'
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f'\n成功转换！输出文件: {output_file}')
        print(f'JSON 文件大小: {len(json.dumps(data, ensure_ascii=False))} 字符')

        # 显示预览
        print('\n数据预览:')
        preview = json.dumps(data, ensure_ascii=False, indent=2)[:500]
        print(preview)
        if len(preview) >= 500:
            print('\n...（预览截断，查看完整数据请打开 JSON 文件）')
    else:
        print('转换失败')
        sys.exit(1)